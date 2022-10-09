import json
import sys

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
import openpyxl
import pandas as pd
import numpy as np


class MainData(QWidget):
    def __init__(self):
        super().__init__()

        self.leftFrame = QFrame()
        self.leftFrame.setFrameShape(QFrame.StyledPanel)

        self.rightFrame = QFrame()
        self.rightFrame.setFrameShape(QFrame.StyledPanel)

        splitter2 = QSplitter(Qt.Horizontal)

        splitter2.addWidget(self.leftFrame)
        splitter2.addWidget(self.rightFrame)
        splitter2.setSizes([22, 22])

        hbox = QHBoxLayout(self)
        hbox.addWidget(splitter2)

        self.setLayout(hbox)
        self.crt()

    def crt(self):
        wb = openpyxl.reader.excel.load_workbook(filename="Daily_Bulletin_07102022.xlsx", data_only=True)
        wb2 = openpyxl.reader.excel.load_workbook(filename="DailyFinancials_EN.xlsx", data_only=True)
        wb.active = 8
        sheet = wb.active
        wb2.active = 0
        sheet2 = wb2.active
        str1 = str(sheet2['AD176'].value)
        str2 = str(sheet['B3'].value)

        self.company_name = QLineEdit("ADX", self.leftFrame)
        self.company_name.setReadOnly(True)
        self.company_name.setGeometry(20, 10, 90, 60)
        self.company_name.setFrame(False)
        font = QFont()
        font.setPointSize(22)
        self.company_name.setFont(font)

        self.company_name2 = QLineEdit(str1, self.leftFrame)
        self.company_name2.setReadOnly(True)
        self.company_name2.setGeometry(20, 70, 150, 60)
        self.company_name2.setFrame(False)
        font = QFont()
        font.setPointSize(22)
        self.company_name2.setFont(font)

        # self.add_btn = QPushButton("Добавить...", self.rightFrame)
        # self.add_btn.setGeometry(QtCore.QRect(10, 160, 200, 40))
        # font = QtGui.QFont()
        # font.setPointSize(10)
        # self.add_btn.setFont(font)
        # self.add_btn.clicked.connect(self.predict)

        self.lineEdit = QLineEdit("DFM", self.rightFrame)
        self.lineEdit.setReadOnly(True)
        self.lineEdit.setGeometry(QtCore.QRect(20, 10, 90, 60))
        self.company_name2.setFrame(False)
        font = QtGui.QFont()
        font.setPointSize(22)
        self.lineEdit.setFont(font)

        self.lineEdit2 = QLineEdit(str2, self.rightFrame)
        self.lineEdit2.setReadOnly(True)
        self.lineEdit2.setGeometry(QtCore.QRect(20, 70, 150, 60))
        self.company_name2.setFrame(False)
        font = QtGui.QFont()
        font.setPointSize(22)
        self.lineEdit2.setFont(font)

        # self.lineEdit_2 = QLineEdit(self.rightFrame)
        # self.lineEdit_2.setGeometry(QtCore.QRect(10, 60, 200, 40))
        # font = QtGui.QFont()
        # font.setPointSize(10)
        # self.lineEdit_2.setFont(font)
        #
        # self.lineEdit_3 = QLineEdit(self.rightFrame)
        # self.lineEdit_3.setGeometry(QtCore.QRect(10, 110, 200, 40))
        # font = QtGui.QFont()
        # font.setPointSize(10)
        # self.lineEdit_3.setFont(font)

    # def predict(self):
    #     wb = openpyxl.reader.excel.load_workbook(filename="Daily_Bulletin_07102022.xlsx", data_only=True)
    #     wb2 = openpyxl.reader.excel.load_workbook(filename="DailyFinancials_EN.xlsx", data_only=True)
    #     wb.active = 8
    #     sheet = wb.active
    #     wb2.active = 1
    #     sheet2 = wb2.active
    #     str1 = "ADX" + sheet2['AD176']
    #     str2 = "DFM" + sheet['B3'].value
    #     self.company_name.setText(str1)
    #     self.lineEdit.setText(str2)
    #     return


    # def save(self):
    #     if self.system.text():
    #         acts.system_name = self.system.text()
    #     dialog = DialogClass()
    #     dialog.setWindowTitle('сохранение')
    #     dialog.setFixedSize(250, 100)
    #     dialog.exec_()


class MainWindow(QMainWindow):
    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent)
        # self.createMenus()
        self.area = MainData()
        self.setCentralWidget(self.area)

    # def createMenus(self):
    #     self.toMenu = self.menuBar().addAction("&Меню", self.menu)
    #
    #     self.exit = self.menuBar().addAction("&Выход", self.close)

    # def save(self):
    #     self.area.save()

    # def menu(self):
    #     self.close()
    #     main.ex = Menu()
    #     main.ex.setGeometry(1000, 1000, 1100, 600)
    #     main.ex.setFixedSize(500, 450)
    #     main.ex.move(QApplication.desktop().screen().rect().center() - main.ex.rect().center())
    #     main.ex.setWindowTitle('TenderHelper')
    #     main.ex.show()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MainWindow()
    ex.setGeometry(300, 500, 650, 500)
    ex.setWindowTitle('TradesCollector')
    ex.move(QApplication.desktop().screen().rect().center() - ex.rect().center())
    ex.show()
    sys.exit(app.exec_())



